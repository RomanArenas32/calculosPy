import openpyxl
import os
import locale
import glob
import datetime

def monedaAFlotante(valor):
    # Verificar si el valor es un número (int o float) y retornarlo tal cual sin aplicar strip()
    if isinstance(valor, (int, float)):
        return valor
    
    # Si el valor no es un número, intentar convertirlo a flotante
    try:
        return locale.atof(valor.strip("$"))
    except ValueError:
        # Si no se puede convertir a flotante, retornar 0 o cualquier otro valor predeterminado
        return 0


def formatearNumero(numero):
    return locale.format_string("%.2f", numero, grouping=True)

inicio = datetime.datetime.now()
print('Fecha y hora de inicio de la exportación: ', inicio)

modoDebug = False
datosExportar = []
locale.setlocale(locale.LC_ALL, 'en_US.utf8')
file_path = "C:\\Users\\PC06\\Desktop\\calculosPy\\ASIGNACIONES CARRERA DOCENTE.xlsx"

# Lista de escuelaNumero para incluir
inclusion_list = [5933, 5998, 6021, 6048, 6075, 6076, 6083, 6127, 6140, 6155, 6433, 6441, 6479, 6972, 7116]

for xlsxFile in glob.glob(file_path):
    print("Procesando ", xlsxFile)
    wb = openpyxl.load_workbook(xlsxFile)
    sheet = wb.active

    # Iterar sobre todas las filas en el archivo Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        escuelaNumero = row[0]
        
        # Check if the escuelaNumero is 'Total general', None, or '(en blanco)', then skip this iteration
        if escuelaNumero in ('Total general', None, '(en blanco)'):
            continue
        
        # Convertir a entero si es un valor válido
        escuelaNumero = int(escuelaNumero)
        
        # Check if the escuelaNumero is in the inclusion list
        if escuelaNumero in inclusion_list:
            sumTotalTransfDirecta = monedaAFlotante(row[6])

            # Resto del código sin cambios...

            datosExportar.append([escuelaNumero, "4-1-01-01-32", "SUBV. BRP TITULO Y MENCION (LEY 20158, ART. 1º)", (sumTotalTransfDirecta)])

excelFile = "resultadosTres.xlsx"
print("Exportando a ", excelFile)

# Crear un nuevo archivo Excel
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'ResultadosExpo'
for row in datosExportar:
    sheet.append(row)

# Formato con 2 decimales para la columna del monto
fmtNumero = openpyxl.styles.NamedStyle(name='custom_number_format', number_format='#,##0.00')
for cell in sheet["D"]:
    cell.style = fmtNumero

# Ajustar el ancho de las columnas
sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 40
sheet.column_dimensions['C'].width = 40
sheet.column_dimensions['D'].width = 16

try:
    wb.save(filename=excelFile)
except PermissionError:
    print(f"No se pudo guardar el archivo '{excelFile}' porque está abierto en otro programa.")

fin = datetime.datetime.now()
print('Fecha y hora de fin de la exportación: ', fin)
print('Tiempo total de ejecución: ', (fin - inicio))
