import openpyxl
import os
import locale
import glob
import datetime

def monedaAFlotante(valor):
    # Verificar si el valor es un número (int o float) y retornarlo tal cual sin aplicar strip()
    if isinstance(valor, (int, float)):
        return valor
    
    # Si el valor no es un número, intentar convertirlo a flotante
    try:
        return locale.atof(valor.strip("$"))
    except ValueError:
        # Si no se puede convertir a flotante, retornar 0 o cualquier otro valor predeterminado
        return 0



inicio = datetime.datetime.now()
print('Fecha y hora de inicio de la exportación: ', inicio)


datosExportar = []
file_path = "C:\\Users\\PC06\\Desktop\\calculosPy\\Sostenedor_Particular_Detalle_Transferido_para_el_Mes_por_Establecimiento_Sostenedor_82733800.xlsx"

# Diccionario para asociar conseptoPago con el número correspondiente
consepto_numero_map = {
    "BONOESCOLAR": "4-1-01-01-20",
    "BONOESCOLARC2": "4-1-01-01-20",
    "PRORETENCION": "4-1-01-01-31",
    "SNED": "4-1-01-01-16",
    "ADECO": "4-1-01-01-70",
    "RFZEDUCATIVO": "4-1-01-01-17",
    "AGUINFPATRIAS": "4-1-01-01-18",
    "AGUINNAVIDAD": "4-1-01-01-19",
    "BONOESPECIAL": "4-1-01-01-24",
    "MANTENIMIENTO": "4-1-01-01-29",
    "BONOVACACIONES": "4-1-01-01-36",
}

tipo_de_subvencion_map = {
    "BONOESCOLAR":	"SUBV. BONO ESCOLARIDAD",
    "BONOESCOLARC2": "SUBV. BONO ESCOLARIDAD",
    "PRORETENCION":	"SUBV. PRO-RETENCION",
    "SNED":	"SUBV. EXC. ACADEMICA",
    "ADECO": "OTRAS SUBVENCIONES ADECO",
    "RFZEDUCATIVO": "SUBV. REFUERZO EDUCATIVO",
    "AGUINFPATRIAS": "SUBV. AGUIN. F. PATRIAS",
    "AGUINNAVIDAD":	"SUBV. AGUINALDO NAVIDAD",
    "BONOESPECIAL":	"BONO TERMINO DE CONFLICTO",
    "MANTENIMIENTO": "SUBV. MANTENIMIENTO",
    "BONOVACACIONES": "BONO VACACIONES DOCENTES Y ASISTENTES ",
}

for xlsxFile in glob.glob(file_path):
    print("Procesando ", xlsxFile)
    wb = openpyxl.load_workbook(xlsxFile)
    sheet = wb.active

    # Iterar sobre todas las filas en el archivo Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        escuelaNumero = int(row[2])  
        conseptoPago = row[10]
        valorMonto = monedaAFlotante(row[8])

        # Obtener el número correspondiente según el conseptoPago
        numero_consepto = consepto_numero_map.get(conseptoPago, "")
        subvencion_consepto = tipo_de_subvencion_map.get(conseptoPago, "")

        datosExportar.append([escuelaNumero, subvencion_consepto, numero_consepto, conseptoPago, valorMonto])

excelFile = "resultadosCuatro.xlsx"
print("Exportando a ", excelFile)

# Crear un nuevo archivo Excel
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'ResultadosExpo'
for row in datosExportar:
    sheet.append(row)

# Formato con 2 decimales para la columna del monto
fmtNumero = openpyxl.styles.NamedStyle(name='custom_number_format', number_format='#,##0.00')
for cell in sheet["D"]:
    cell.style = fmtNumero

# Ajustar el ancho de las columnas
sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 40
sheet.column_dimensions['C'].width = 40
sheet.column_dimensions['D'].width = 16

try:
    wb.save(filename=excelFile)
except PermissionError:
    print(f"No se pudo guardar el archivo '{excelFile}' porque está abierto en otro programa.")

fin = datetime.datetime.now()
print('Fecha y hora de fin de la exportación: ', fin)
print('Tiempo total de ejecución: ', (fin - inicio))
