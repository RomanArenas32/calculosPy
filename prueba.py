import openpyxl
import os
import locale
import glob
import datetime

def monedaAFlotante(valor):
    if isinstance(valor, (int, float)):
        return valor
    try:
        return locale.atof(valor.strip("$"))
    except ValueError: 
        return 0  

inicio = datetime.datetime.now()
print('Fecha y hora de inicio de la exportación: ', inicio)


datosExportar = []
file_path = "C:\\Users\\PC06\\Desktop\\calculosPy\\xls_python.xlsx"

for xlsxFile in glob.glob(file_path):
    print("Procesando ", xlsxFile)
    wb = openpyxl.load_workbook(xlsxFile)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        escuelaNumero = int(row[0])
        valSubvencionBase = monedaAFlotante(row[1])
        valInternado = monedaAFlotante(row[2])
        valIncrementoZona = monedaAFlotante(row[3])
        valIncrementoRuralidad = monedaAFlotante(row[4])
        valDiferenciaPisoRural = monedaAFlotante(row[5])
        valAporteGratuidad = monedaAFlotante(row[8])
        valMontoLey19410 = monedaAFlotante(row[11])
        valMontoLey19464 = monedaAFlotante(row[12])
        valMontoLey19464NoDoc = monedaAFlotante(row[13])
        valDesempenoDificilDocentes = monedaAFlotante(row[14])
        valDesempenoDificilAsistentes = monedaAFlotante(row[15])
        valProfesorEncargado = monedaAFlotante(row[16])
        valMontoPie = monedaAFlotante(row[29])
        valTotalDescuentos = monedaAFlotante(row[24])
        ReliquidaciónPeriodoMarzoAMayo = monedaAFlotante(row[25])


        datosExportar.append([escuelaNumero, "4-1-1-01-10", "Subvención general", (valSubvencionBase+valIncrementoZona-valMontoPie-valTotalDescuentos+ReliquidaciónPeriodoMarzoAMayo)])
        datosExportar.append([escuelaNumero, "4-1-1-01-11", "Subvención internado", (valInternado)])
        datosExportar.append([escuelaNumero, "4-1-1-01-12", "Subvención educación especial", (valMontoPie)])
        datosExportar.append([escuelaNumero, "4-1-1-01-13", "Subvención desempeno difícil", (valDesempenoDificilAsistentes + valDesempenoDificilDocentes)])
        datosExportar.append([escuelaNumero, "4-1-1-01-14", "Subvención Ley 19.410", (valMontoLey19410)])
        datosExportar.append([escuelaNumero, "4-1-1-01-21", "Subvención personal no docente", (valMontoLey19464 + valMontoLey19464NoDoc)])
        datosExportar.append([escuelaNumero, "4-1-1-01-30", "Subvención prof. encargado esc. rurales", (valProfesorEncargado)])
        datosExportar.append([escuelaNumero, "4-1-1-01-33", "Subvención ruralidad", (valIncrementoRuralidad + valDiferenciaPisoRural)])
        datosExportar.append([escuelaNumero, "4-1-1-01-89", "Aporte gratuidad ley inclusión", (valAporteGratuidad)])

excelFile = "resultados.xlsx"
print("Exportando a ", excelFile)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'ResultadosExpo'
for row in datosExportar:
    sheet.append(row)

fmtNumero = openpyxl.styles.NamedStyle(name='custom_number_format', number_format='#,##0.00')
for cell in sheet["D"]:
    cell.style = fmtNumero

sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 40
sheet.column_dimensions['C'].width = 40
sheet.column_dimensions['D'].width = 16

try:
    wb.save(filename=excelFile)
except PermissionError:
    print(f"No se pudo guardar el archivo '{excelFile}' porque está abierto en otro programa.")

fin = datetime.datetime.now()
print('Fecha y hora de fin de la exportación: ', fin)
print('Tiempo total de ejecución: ', (fin - inicio))
